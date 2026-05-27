"""Build and verify Excel workbooks for SpreadsheetBench-style tasks."""

from __future__ import annotations

import json
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook, load_workbook
except ImportError:  # pragma: no cover - optional until installed
    Workbook = None  # type: ignore
    load_workbook = None  # type: ignore


def require_openpyxl() -> None:
    if Workbook is None:
        raise ImportError("Install openpyxl: pip install skillopt[spreadsheet]")


def create_workbook(path: Path, sheet_data: list[list[Any]], sheet_name: str = "Sheet1") -> None:
    require_openpyxl()
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    for r_idx, row in enumerate(sheet_data, start=1):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def workbook_preview(path: Path, max_rows: int = 20, max_cols: int = 10) -> str:
    require_openpyxl()
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    lines = [f"Sheet: {ws.title}"]
    for r in range(1, min(max_rows, ws.max_row or 1) + 1):
        row_vals = []
        for c in range(1, min(max_cols, ws.max_column or 1) + 1):
            val = ws.cell(row=r, column=c).value
            col = _col_letter(c)
            row_vals.append(f"{col}{r}={val!r}")
        lines.append(" | ".join(row_vals))
    return "\n".join(lines)


def apply_writes(path: Path, writes: list[dict[str, Any]]) -> list[str]:
    """Apply cell writes. Returns log lines."""
    require_openpyxl()
    wb = load_workbook(path)
    ws = wb.active
    log: list[str] = []
    for w in writes:
        cell = w.get("cell") or w.get("address")
        value = w.get("value")
        if not cell:
            continue
        ws[cell] = value
        log.append(f"write {cell}={value!r}")
    wb.save(path)
    return log


def verify_against_answer_workbook(
    output_path: Path,
    answer_path: Path,
    cells: list[str],
    tolerance: float = 0.01,
) -> tuple[float, dict]:
    """Compare output workbook cells against official SpreadsheetBench answer file."""
    require_openpyxl()
    if not cells:
        return 0.0, {"error": "no answer cells specified"}

    out_wb = load_workbook(output_path, data_only=True)
    ans_wb = load_workbook(answer_path, data_only=True)
    out_ws = out_wb.active
    ans_ws = ans_wb.active

    correct = 0
    details: dict[str, dict] = {}
    for cell in cells:
        actual = out_ws[cell].value
        expected = ans_ws[cell].value
        ok = _values_match(actual, expected, tolerance)
        details[cell] = {
            "expected": _serialize_cell_value(expected),
            "actual": _serialize_cell_value(actual),
            "ok": ok,
        }
        if ok:
            correct += 1

    return correct / len(cells), details


def expected_cells_from_answer_workbook(
    answer_path: Path,
    cells: list[str],
) -> dict[str, Any]:
    """Build expected_cells map from answer xlsx (for mock YAML export)."""
    require_openpyxl()
    wb = load_workbook(answer_path, data_only=True)
    ws = wb.active
    return {cell: ws[cell].value for cell in cells}


def verify_cells(path: Path, expected_cells: dict[str, Any], tolerance: float = 0.01) -> tuple[float, dict]:
    require_openpyxl()
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    if not expected_cells:
        return 0.0, {}

    correct = 0
    details: dict[str, dict] = {}
    for cell, expected in expected_cells.items():
        actual = ws[cell].value
        ok = _values_match(actual, expected, tolerance)
        details[cell] = {
            "expected": _serialize_cell_value(expected),
            "actual": _serialize_cell_value(actual),
            "ok": ok,
        }
        if ok:
            correct += 1

    score = correct / len(expected_cells)
    return score, details


def parse_writes_json(text: str) -> list[dict[str, Any]]:
    text = text.strip()
    if text.startswith("```"):
        text = text.split("\n", 1)[-1].rsplit("```", 1)[0].strip()
    data = json.loads(text)
    if isinstance(data, dict) and "writes" in data:
        return list(data["writes"])
    if isinstance(data, list):
        return data
    raise ValueError("Expected JSON with 'writes' array")


def _serialize_cell_value(value: Any) -> Any:
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    return value


def _values_match(actual: Any, expected: Any, tolerance: float) -> bool:
    if actual is None and expected is None:
        return True
    if actual is None or expected is None:
        return False
    try:
        if isinstance(expected, (int, float)) and isinstance(actual, (int, float)):
            return abs(float(actual) - float(expected)) <= tolerance
    except (TypeError, ValueError):
        pass
    return str(actual).strip().lower() == str(expected).strip().lower()


def _col_letter(col: int) -> str:
    letters = ""
    while col:
        col, rem = divmod(col - 1, 26)
        letters = chr(65 + rem) + letters
    return letters
